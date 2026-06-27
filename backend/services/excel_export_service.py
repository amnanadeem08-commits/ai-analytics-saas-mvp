from __future__ import annotations

import io
from typing import Any

import pandas as pd
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
SECTION_FILL = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _safe_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [{key: str(value) if isinstance(value, (dict, list)) else value for key, value in row.items()} for row in rows]


def _coerce_float(value: Any) -> float | None:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except Exception:
        return None


def _fallback_kpis(raw_df: pd.DataFrame) -> list[dict[str, Any]]:
    total_cells = max(int(raw_df.shape[0] * raw_df.shape[1]), 1)
    missing_rate = round(float(raw_df.isna().sum().sum()) / total_cells * 100, 2)
    numeric = raw_df.select_dtypes(include="number")
    kpis = [
        {"label": "Row Count", "value": int(len(raw_df)), "delta_percentage": None, "status": "neutral"},
        {"label": "Column Count", "value": int(len(raw_df.columns)), "delta_percentage": None, "status": "neutral"},
        {"label": "Numeric Column Count", "value": int(len(numeric.columns)), "delta_percentage": None, "status": "neutral"},
        {"label": "Missing Value Rate %", "value": missing_rate, "delta_percentage": None, "status": "warning" if missing_rate > 10 else "positive"},
    ]
    for column in numeric.columns[:2]:
        kpis.append(
            {
                "label": f"Average {column.replace('_', ' ').title()}",
                "value": round(float(pd.to_numeric(raw_df[column], errors="coerce").mean()), 4),
                "delta_percentage": None,
                "status": "neutral",
            }
        )
    return kpis


def _chart_rows_from_spec(chart: dict[str, Any]) -> tuple[list[str], list[float]]:
    traces = chart.get("plotly", {}).get("data", [])
    if not traces:
        return [], []
    trace = traces[0]
    labels = trace.get("x", [])
    values = trace.get("y", [])
    if chart.get("chart_type") == "pie":
        labels = trace.get("labels", [])
        values = trace.get("values", [])
    clean_labels: list[str] = []
    clean_values: list[float] = []
    for label, value in zip(labels, values):
        numeric = _coerce_float(value)
        if numeric is None:
            continue
        clean_labels.append(str(label))
        clean_values.append(numeric)
    return clean_labels[:20], clean_values[:20]


def _fallback_chart_series(raw_df: pd.DataFrame) -> tuple[str, list[str], list[float], str]:
    numeric_columns = raw_df.select_dtypes(include="number").columns.tolist()
    if numeric_columns:
        metric = numeric_columns[0]
        date_candidates = [col for col in raw_df.columns if "date" in str(col).lower() or "time" in str(col).lower()]
        if date_candidates:
            date_col = date_candidates[0]
            work = raw_df[[date_col, metric]].copy()
            work[date_col] = pd.to_datetime(work[date_col], errors="coerce")
            work[metric] = pd.to_numeric(work[metric], errors="coerce")
            work = work.dropna()
            if not work.empty:
                work["_period"] = work[date_col].dt.to_period("M").astype(str)
                grouped = work.groupby("_period")[metric].mean().sort_index()
                return f"Average {metric} by Month", grouped.index.astype(str).tolist(), grouped.values.astype(float).tolist(), "line"
        categorical = [col for col in raw_df.columns if col not in numeric_columns][:1]
        if categorical:
            dim = categorical[0]
            grouped = (
                raw_df.assign(_metric=pd.to_numeric(raw_df[metric], errors="coerce"))
                .groupby(dim, dropna=False)["_metric"]
                .sum()
                .sort_values(ascending=False)
                .head(12)
            )
            return f"Total {metric} by {dim}", [str(i) for i in grouped.index], grouped.values.astype(float).tolist(), "bar"
    counts = raw_df.columns.to_series().head(10).value_counts()
    return "Column Frequency", [str(i) for i in counts.index], counts.values.astype(float).tolist(), "bar"


def _apply_table_format(ws, start_row: int, end_row: int, end_col: int) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=end_col):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")


def _autofit_columns(ws) -> None:
    for col_idx, column in enumerate(ws.iter_cols(min_col=1, max_col=ws.max_column), start=1):
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 12), 48)


def _write_kpis(ws, row: int, kpis: list[dict[str, Any]]) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value="KPI Overview")
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.cell(row=row, column=1).font = HEADER_FONT
    row += 1

    headers = ["KPI", "Value", "Delta %", "Status"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    row += 1

    for card in kpis[:16]:
        ws.cell(row=row, column=1, value=card.get("label", ""))
        ws.cell(row=row, column=2, value=card.get("value", ""))
        ws.cell(row=row, column=3, value=_coerce_float(card.get("delta_percentage")))
        ws.cell(row=row, column=4, value=str(card.get("status", "")).title())
        row += 1
    _apply_table_format(ws, row - len(kpis[:16]) - 1, row - 1, 4)
    return row + 1


def _write_insights(ws, row: int, executive: dict[str, Any]) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value="Executive Insights")
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.cell(row=row, column=1).font = HEADER_FONT
    row += 1
    insight_rows = [
        ("What happened", executive.get("insight", "")),
        ("Why it happened", executive.get("reason", "")),
        ("What to do", executive.get("action", "")),
    ]
    for title, text in insight_rows:
        ws.cell(row=row, column=1, value=title).font = Font(bold=True)
        ws.cell(row=row, column=2, value=text)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1
    _apply_table_format(ws, row - len(insight_rows), row - 1, 4)
    return row + 1


def _add_chart(sheet, start_row: int, title: str, labels: list[str], values: list[float], chart_type: str) -> int:
    if len(labels) < 2 or len(values) < 2:
        return start_row
    sheet.cell(row=start_row, column=1, value="Label").font = Font(bold=True)
    sheet.cell(row=start_row, column=2, value=title).font = Font(bold=True)
    for idx, (label, value) in enumerate(zip(labels, values), start=start_row + 1):
        sheet.cell(row=idx, column=1, value=label)
        sheet.cell(row=idx, column=2, value=value)
    data_ref = Reference(sheet, min_col=2, min_row=start_row, max_row=start_row + len(values))
    cats_ref = Reference(sheet, min_col=1, min_row=start_row + 1, max_row=start_row + len(values))
    if chart_type == "line":
        chart = LineChart()
    elif chart_type == "pie":
        chart = PieChart()
    else:
        chart = BarChart()
    chart.title = title[:120]
    chart.height = 7
    chart.width = 12
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    sheet.add_chart(chart, f"E{start_row}")
    return start_row + max(len(values) + 2, 24)



def _storyboard_section(payload: dict[str, Any], section_id: str) -> dict[str, Any]:
    return next((section for section in payload.get("sections", []) if section.get("section_id") == section_id), {})


def _write_storyboard_summary(wb, report: dict[str, Any]) -> None:
    storyboard = report.get("executive_storyboard", {})
    ws = wb.create_sheet("Storyboard Summary", 0)
    ws.merge_cells("A1:E1")
    ws["A1"] = "Executive Storyboard Summary"
    ws["A1"].font = Font(size=16, bold=True, color="1F4E78")
    ws["A2"] = f"Dataset: {report.get('dataset_id', '')}"
    row = 4

    summary = _storyboard_section(storyboard, "executive_summary").get("content", {})
    readiness = summary.get("dataset_readiness", {})
    ws.cell(row=row, column=1, value="Executive Summary").fill = SECTION_FILL
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    for label, value in [
        ("Dataset Readiness", f"{readiness.get('score', 0)}/100"),
        ("Overall Business Health", f"{summary.get('overall_business_health', 0)}/100"),
        ("Executive Summary", summary.get("executive_summary", "")),
        ("Top Opportunity", summary.get("top_opportunity", "")),
        ("Biggest Risk", summary.get("biggest_risk", "")),
    ]:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        row += 1
    row += 1

    kpis = _storyboard_section(storyboard, "kpi_overview").get("kpis", [])
    ws.cell(row=row, column=1, value="KPI Summary").fill = SECTION_FILL
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    headers = ["KPI", "Value", "Status", "Context"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=row, column=idx, value=header).fill = HEADER_FILL
        ws.cell(row=row, column=idx).font = HEADER_FONT
    row += 1
    for card in kpis[:12]:
        ws.cell(row=row, column=1, value=card.get("label", ""))
        ws.cell(row=row, column=2, value=card.get("value", ""))
        ws.cell(row=row, column=3, value=card.get("status", ""))
        ws.cell(row=row, column=4, value=card.get("business_context", card.get("description", "")))
        row += 1
    row += 1

    cards = _storyboard_section(storyboard, "ai_business_insights").get("cards", [])
    ws.cell(row=row, column=1, value="AI Business Insights").fill = SECTION_FILL
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    headers = ["Type", "Title", "Evidence", "Impact", "Recommendation"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=row, column=idx, value=header).fill = HEADER_FILL
        ws.cell(row=row, column=idx).font = HEADER_FONT
    row += 1
    for card in cards:
        ws.cell(row=row, column=1, value=card.get("type", ""))
        ws.cell(row=row, column=2, value=card.get("title", ""))
        ws.cell(row=row, column=3, value=card.get("supporting_evidence", ""))
        ws.cell(row=row, column=4, value=card.get("expected_business_impact", ""))
        ws.cell(row=row, column=5, value=card.get("executive_recommendation", ""))
        row += 1
    row += 1

    recs = _storyboard_section(storyboard, "executive_recommendations").get("recommendations", [])
    ws.cell(row=row, column=1, value="Executive Recommendations").fill = SECTION_FILL
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    headers = ["Priority", "Business Value", "Difficulty", "Expected Impact", "Recommendation"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=row, column=idx, value=header).fill = HEADER_FILL
        ws.cell(row=row, column=idx).font = HEADER_FONT
    row += 1
    for rec in recs:
        ws.cell(row=row, column=1, value=rec.get("priority", ""))
        ws.cell(row=row, column=2, value=rec.get("business_value", ""))
        ws.cell(row=row, column=3, value=rec.get("difficulty", ""))
        ws.cell(row=row, column=4, value=rec.get("expected_impact", ""))
        ws.cell(row=row, column=5, value=rec.get("recommendation", ""))
        row += 1
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
    _autofit_columns(ws)

def build_executive_excel(
    report: dict[str, Any],
    raw_df: pd.DataFrame,
    summary: dict[str, Any],
    chart_ids: list[str] | None = None,
    package: str = "executive",
) -> bytes:
    buffer = io.BytesIO()
    selected = set(chart_ids or [])
    charts = [chart for chart in report.get("chart_specs", []) if not selected or chart.get("chart_id") in selected]
    kpis = report.get("kpi_cards", []) or _fallback_kpis(raw_df)
    executive = report.get("executive_summary", {})

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        wb = writer.book
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            wb.remove(wb["Sheet"])
        if package == "storyboard":
            _write_storyboard_summary(wb, report)
        summary_ws = wb.create_sheet("Dashboard Summary")
        summary_ws.merge_cells("A1:D1")
        summary_ws["A1"] = report.get("branding", {}).get("report_title", "Executive Dashboard Summary")
        summary_ws["A1"].font = Font(size=16, bold=True, color="1F4E78")
        summary_ws["A2"] = f"Dataset: {report.get('dataset_id', '')}"
        summary_ws["A3"] = f"Package: {package.title()}"

        next_row = _write_kpis(summary_ws, 5, kpis)
        next_row = _write_insights(summary_ws, next_row, executive)

        visual_ws = wb.create_sheet("Visual Dashboard")
        visual_ws["A1"] = "Executive Visual Dashboard"
        visual_ws["A1"].font = Font(size=14, bold=True, color="1F4E78")
        chart_row = 3
        chart_added = 0
        for chart in charts[:6]:
            labels, values = _chart_rows_from_spec(chart)
            chart_type = "line" if chart.get("chart_type") in {"line"} else "pie" if chart.get("chart_type") == "pie" else "bar"
            new_row = _add_chart(visual_ws, chart_row, chart.get("title", "Chart"), labels, values, chart_type)
            if new_row != chart_row:
                chart_added += 1
                chart_row = new_row
        if chart_added == 0 and not selected:
            title, labels, values, fallback_type = _fallback_chart_series(raw_df)
            _add_chart(visual_ws, chart_row, title, labels, values, fallback_type)

        storyboard_items = report.get("storyboard_items") or report.get("storyboard") or []
        if storyboard_items:
            pd.DataFrame(_safe_rows(storyboard_items)).to_excel(writer, sheet_name="Storyboard", index=False)

        raw_df.to_excel(writer, sheet_name="Raw Data", index=False)
        missing_by_column = summary.get("missing_values_by_column", {})
        dtypes = summary.get("dtypes", {})
        schema_rows = [
            {
                "column": column,
                "dtype": dtypes.get(column, str(raw_df[column].dtype)),
                "missing_count": int(missing_by_column.get(column, 0)),
                "unique_count": int(raw_df[column].nunique(dropna=True)),
                "completeness_pct": round((1 - int(missing_by_column.get(column, 0)) / max(len(raw_df), 1)) * 100, 2),
            }
            for column in raw_df.columns
        ]
        pd.DataFrame(schema_rows).to_excel(writer, sheet_name="Column Schema", index=False)

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            _autofit_columns(ws)

    return buffer.getvalue()
