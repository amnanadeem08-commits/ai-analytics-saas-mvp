from io import BytesIO
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook

from backend.services.excel_export_service import build_executive_excel
from backend.services.pdf_export_service import build_executive_pdf
from backend.services.ppt_export_service import build_executive_pptx



def _report_payload() -> dict:
    chart = {
        "chart_id": "chart_revenue",
        "title": "Revenue by Segment",
        "chart_type": "bar",
        "section": "comparisons",
        "columns": ["segment", "revenue"],
        "plotly": {
            "data": [{"type": "bar", "x": ["A", "B"], "y": [100, 180]}],
            "layout": {"title": "Revenue by Segment"},
        },
        "metadata": {"subtitle": "Revenue distribution"},
    }
    return {
        "dataset_id": "ds_visual_integrity",
        "branding": {"company_name": "AI Analytics", "report_title": "Executive Report", "primary_color": "#118DFF", "accent_color": "#E66C37"},
        "overview": {"row_count": 2, "column_count": 2},
        "kpi_cards": [{"kpi_id": "kpi_revenue", "label": "Total Revenue", "value": 280, "formatted_value": "280"}],
        "chart_specs": [chart],
        "chart_count": 1,
        "executive_summary": {
            "insight": "Revenue is concentrated in segment B.",
            "reason": "Segment B contributes most total revenue.",
            "action": "Prioritize segment B growth with retention safeguards.",
            "recommendations": [{"recommendation": "Scale successful segment B campaigns.", "reason": "Strong revenue signal", "expected_impact": "Revenue lift"}],
        },
        "business_story": {"business_story": "Segment B drives business outcomes."},
        "data_quality_score": {"score": 95, "grade": "A", "completeness_pct": 100, "duplicate_pct": 0, "explanation": "High quality"},
    }



def test_executive_excel_contains_visual_dashboard_and_pivot_charts():
    payload = _report_payload()
    raw_df = pd.DataFrame({"segment": ["A", "B"], "revenue": [100, 180]})

    data = build_executive_excel(payload, raw_df=raw_df, summary={"missing_values_by_column": {}, "dtypes": {}}, package="executive")
    workbook = load_workbook(BytesIO(data))

    assert "Visual Dashboard" in workbook.sheetnames
    assert "Pivot Tables" in workbook.sheetnames
    assert len(workbook["Visual Dashboard"]._images) >= 1
    assert len(workbook["Pivot Tables"]._charts) >= 1



def test_executive_pptx_contains_embedded_visual_media():
    payload = _report_payload()
    pptx_bytes = build_executive_pptx(payload, package="executive")

    with ZipFile(BytesIO(pptx_bytes)) as archive:
        assert any(name.startswith("ppt/media/image") for name in archive.namelist())



def test_executive_pdf_contains_embedded_chart_images():
    payload = _report_payload()
    pdf_bytes = build_executive_pdf(payload, package="executive")

    assert pdf_bytes.startswith(b"%PDF")
    assert b"/Subtype /Image" in pdf_bytes or b"/Subtype/Image" in pdf_bytes
