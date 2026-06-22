from pathlib import Path
from io import BytesIO
import logging
import shutil
from uuid import uuid4

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from backend.core.config import ensure_data_directories, settings
from backend.core.theme_manager import theme_manager
from backend.main import app
from frontend.components.storyboard_session import add_storyboard_entry


def test_upload_persists_dataset_record_and_status(monkeypatch, caplog):
    test_root = Path("data") / "test_runs" / uuid4().hex
    monkeypatch.setattr(settings, "DATA_DIR", test_root)
    monkeypatch.setattr(settings, "UPLOADS_DIR", test_root / "uploads")
    monkeypatch.setattr(settings, "PROCESSED_DIR", test_root / "processed")
    monkeypatch.setattr(settings, "METADATA_DIR", test_root / "metadata")
    monkeypatch.setattr(settings, "DATASETS_DIR", test_root / "datasets")
    monkeypatch.setattr(settings, "SAMPLES_DIR", test_root / "samples")
    monkeypatch.setattr(settings, "BRAND_ASSETS_DIR", test_root / "branding")
    monkeypatch.setattr(settings, "DATASETS_METADATA_FILE", test_root / "metadata" / "datasets.json")
    monkeypatch.setattr(settings, "BRANDING_FILE", test_root / "metadata" / "branding.json")
    monkeypatch.setattr(settings, "SQL_QUERIES_FILE", test_root / "metadata" / "sql_queries.json")
    monkeypatch.setattr(settings, "THEME_STATE_FILE", test_root / "metadata" / "theme_state.json")
    ensure_data_directories()
    theme_manager.set_active("power_bi_professional")

    try:
        client = TestClient(app)
        response = client.post(
            "/upload",
            files={"file": ("sales.csv", b"region,sales\nNorth,100\nSouth,150\n", "text/csv")},
        )

        assert response.status_code == 200
        dataset_id = response.json()["dataset_id"]

        status = client.get(f"/datasets/{dataset_id}/status")
        assert status.status_code == 200
        assert status.json() == {
            "dataset_id": dataset_id,
            "status": "ready",
            "row_count": 2,
            "column_count": 2,
            "error_message": None,
        }

        meta_path = Path(settings.DATASETS_DIR) / dataset_id / "meta.json"
        assert meta_path.exists()

        overview = client.get(f"/datasets/{dataset_id}/overview")
        assert overview.status_code == 200
        overview_body = overview.json()
        assert overview_body["row_count"] == 2
        assert overview_body["column_count"] == 2
        assert overview_body["column_groups"]["numeric"] == ["sales"]
        assert overview_body["column_groups"]["categorical"] == ["region"]
        assert overview_body["missing_summary"]["completeness_pct"] == 100.0
        assert overview_body["column_schema"][0]["name"] == "region"
        assert overview_body["preview"][0] == {"region": "North", "sales": 100}

        dashboard = client.get(f"/analytics/{dataset_id}/dashboard")
        assert dashboard.status_code == 200
        dashboard_body = dashboard.json()
        assert dashboard_body["theme"]["name"] == "power_bi_professional"
        assert dashboard_body["branding"]["company_name"] == "AI Analytics"
        assert dashboard_body["kpi_cards"]
        assert "trend_arrow" in dashboard_body["kpi_cards"][0]
        assert "status_color" in dashboard_body["kpi_cards"][0]
        business_kpi = next(card for card in dashboard_body["kpi_cards"] if card["kpi_id"] == "total_sales")
        assert business_kpi["sparkline"]
        assert business_kpi["reason"]
        assert business_kpi["recommended_action"]
        assert business_kpi["expected_impact"]
        assert business_kpi["evidence"]["top_segment"]["segment"] == "South"
        assert dashboard_body["business_metrics"]["primary_metric"] == "sales"
        assert dashboard_body["domain_intelligence"]["detection"]["domain"] == "Sales"
        assert dashboard_body["regional_analytics"]["available"] is True
        assert dashboard_body["analysis_guardrails"]["supports"]["comparison_analysis"] is True
        assert dashboard_body["analysis_guardrails"]["supports"]["time_intelligence"] is False
        assert dashboard_body["data_quality_score"]["grade"]
        assert dashboard_body["suggested_questions"]
        assert any(chart["section"] == "geographic" for chart in dashboard_body["chart_specs"])
        assert dashboard_body["business_metrics"]["segment_leader"]["dimension"] == "region"
        assert dashboard_body["business_metrics"]["segment_leader"]["metric"] == "sales"
        assert dashboard_body["business_metrics"]["segment_leader"]["segment"] == "South"
        assert dashboard_body["business_metrics"]["segment_leader"]["value"] == 150.0
        assert dashboard_body["business_metrics"]["segment_leader"]["aggregation"] == "sum"
        assert dashboard_body["business_metrics"]["segment_leader"]["business_relevance"] == "high"
        assert dashboard_body["chart_specs"]
        first_chart = dashboard_body["chart_specs"][0]
        assert first_chart["plotly"]["data"]
        assert first_chart["plotly"]["layout"]
        assert first_chart["metadata"]["theme"] == "power_bi_professional"
        assert first_chart["metadata"]["drilldown_ready"] is True
        assert first_chart["chart_type"] in {"histogram", "bar", "pie", "line", "scatter"}
        assert "top_categories" in dashboard_body
        assert "time_trends" in dashboard_body

        insights = client.get(f"/insights/{dataset_id}")
        assert insights.status_code == 200
        insight_body = insights.json()
        assert insight_body["insights"]
        assert insight_body["executive_summary"]["insight"]
        assert insight_body["executive_summary"]["reason"]
        assert insight_body["executive_summary"]["action"]
        assert insight_body["executive_summary"]["evidence"]
        assert insight_body["executive_summary"]["confidence"] == "low"
        assert insight_body["executive_summary"]["key_findings"]
        assert insight_body["executive_summary"]["opportunities"]
        assert insight_body["executive_summary"]["recommendations"]
        assert insight_body["executive_summary"]["action_plan"]
        assert insight_body["executive_summary"]["ceo_insights"][0]["what_happened"]
        assert insight_body["executive_summary"]["ceo_insights"][0]["evidence"]
        decision_block = insight_body["executive_summary"]["decision_framework"][0]
        assert decision_block["framework"] == "what_why_action"
        assert decision_block["what_happened"]
        assert decision_block["why_it_happened"]
        assert decision_block["what_to_do"]
        assert decision_block["expected_impact"]
        assert decision_block["evidence"]

        decision_framework = client.get(f"/insights/{dataset_id}/decision-framework")
        assert decision_framework.status_code == 200
        decision_body = decision_framework.json()
        assert decision_body["framework"] == "what_why_action"
        assert decision_body["decision_blocks"][0]["what_happened"]

        analyst = client.post(
            f"/insights/{dataset_id}/ask",
            json={"question": "Which region has the highest sales?"},
        )
        assert analyst.status_code == 200
        analyst_body = analyst.json()
        assert analyst_body["answer"] == "The top region by total sales is South with 150."
        assert analyst_body["analyst"]["intent"] == "top"
        assert analyst_body["analyst"]["metric_column"] == "sales"
        assert analyst_body["analyst"]["dimension_column"] == "region"
        assert analyst_body["analyst"]["render_mode"] == "ranked_table"
        assert analyst_body["analyst"]["rows"][0]["label"] == "South"

        executive_analyst = client.post(
            f"/insights/{dataset_id}/ask",
            json={"question": "What should management prioritize?"},
        )
        assert executive_analyst.status_code == 200
        executive_analyst_body = executive_analyst.json()
        assert executive_analyst_body["analyst"]["intent"] == "executive_decision"
        assert executive_analyst_body["supporting_data"]["selected_block"]["evidence"]

        filtered = client.post(
            f"/analytics/{dataset_id}/dashboard/filter",
            json={"filters": {"region": {"values": ["South"]}}},
        )
        assert filtered.status_code == 200
        filtered_body = filtered.json()
        assert filtered_body["filtered"] is True
        assert filtered_body["original_row_count"] == 2
        assert filtered_body["filtered_row_count"] == 1
        assert filtered_body["business_metrics"]["segment_leader"]["segment"] == "South"

        visual_schema = client.get(f"/visual-builder/{dataset_id}/schema")
        assert visual_schema.status_code == 200
        visual_schema_body = visual_schema.json()
        assert visual_schema_body["suggested_defaults"]["dimension"] == "region"
        assert visual_schema_body["suggested_defaults"]["measure"] == "sales"
        semantic_roles = {field["name"]: field["semantic_role"] for field in visual_schema_body["semantic_layer"]}
        assert semantic_roles["region"] == "geography_column"
        assert semantic_roles["sales"] == "revenue_currency_column"
        assert visual_schema_body["recommended_visuals"]
        recommended_titles = {visual["title"] for visual in visual_schema_body["recommended_visuals"]}
        assert "Total Sales" in recommended_titles
        assert "Sales by Region" in recommended_titles

        visual = client.post(
            f"/visual-builder/{dataset_id}/render",
            json={
                "chart_type": "horizontal_bar",
                "dimension": "region",
                "measure": "sales",
                "aggregation": "sum",
                "sort": "descending",
                "number_format": "Currency",
                "data_labels": True,
            },
        )
        assert visual.status_code == 200
        visual_body = visual.json()
        assert visual_body["chart"]["chart_type"] == "horizontal_bar"
        assert visual_body["chart"]["plotly"]["data"]
        assert visual_body["applied_spec"]["dimension"] == "region"
        assert visual_body["applied_spec"]["number_format"] == "Currency"
        assert visual_body["chart"]["metadata"]["short_ai_insight"]

        register_visual = client.post(
            f"/visual-builder/{dataset_id}/register",
            json=visual_body["chart"],
        )
        assert register_visual.status_code == 200
        register_visual_body = register_visual.json()
        assert register_visual_body["registered"] is True
        assert register_visual_body["chart"]["chart_id"] == visual_body["chart"]["chart_id"]

        register_visual_duplicate = client.post(
            f"/visual-builder/{dataset_id}/register",
            json=visual_body["chart"],
        )
        assert register_visual_duplicate.status_code == 200
        assert register_visual_duplicate.json()["registered"] is False

        bar_visual = client.post(
            f"/visual-builder/{dataset_id}/render",
            json={
                "chart_type": "bar",
                "dimension": "region",
                "measure": "sales",
                "aggregation": "sum",
                "sort": "descending",
                "data_labels": True,
            },
        )
        assert bar_visual.status_code == 200
        bar_visual_body = bar_visual.json()
        assert bar_visual_body["chart"]["chart_type"] == "bar"
        assert bar_visual_body["chart"]["plotly"]["data"]

        register_bar_visual = client.post(
            f"/visual-builder/{dataset_id}/register",
            json=bar_visual_body["chart"],
        )
        assert register_bar_visual.status_code == 200
        assert register_bar_visual.json()["registered"] is True

        table_visual = client.post(
            f"/visual-builder/{dataset_id}/render",
            json={
                "chart_type": "table",
                "dimension": "region",
                "measure": "sales",
                "aggregation": "sum",
                "sort": "descending",
                "data_labels": True,
            },
        )
        assert table_visual.status_code == 200
        table_visual_body = table_visual.json()
        assert table_visual_body["chart"]["chart_type"] == "table"
        assert table_visual_body["chart"]["plotly"]["data"]

        register_table_visual = client.post(
            f"/visual-builder/{dataset_id}/register",
            json=table_visual_body["chart"],
        )
        assert register_table_visual.status_code == 200
        assert register_table_visual.json()["registered"] is True

        studio_chart_ids = [
            visual_body["chart"]["chart_id"],
            bar_visual_body["chart"]["chart_id"],
            table_visual_body["chart"]["chart_id"],
        ]
        fallback_text = b"Chart image could not be rendered"
        image_marker = b"/Subtype /Image"

        def _assert_pdf_export_for_chart_ids(chart_ids: list[str], minimum_bytes: int) -> bytes:
            caplog.clear()
            caplog.set_level(logging.ERROR, logger="backend.services.export_render_service")
            export_response = client.get(
                f"/report/{dataset_id}/export",
                params=[("format", "pdf"), ("package", "board")] + [("chart_ids", chart_id) for chart_id in chart_ids],
            )
            assert export_response.status_code == 200, (
                f"Studio PDF export did not return HTTP 200 for chart_ids={chart_ids}.\n"
                f"render exceptions:\n{caplog.text or 'No export_render_service exceptions captured.'}"
            )
            assert export_response.content.startswith(b"%PDF")
            assert len(export_response.content) > minimum_bytes, (
                f"Studio PDF export appears too small for chart_ids={chart_ids}: {len(export_response.content)} bytes.\n"
                f"render exceptions:\n{caplog.text or 'No export_render_service exceptions captured.'}"
            )
            assert fallback_text not in export_response.content, (
                "Studio chart export hit fallback rendering in PDF.\n"
                f"chart_ids={chart_ids}\n"
                f"render exceptions:\n{caplog.text or 'No export_render_service exceptions captured.'}"
            )
            return export_response.content

        studio_pdf_content = _assert_pdf_export_for_chart_ids(studio_chart_ids, minimum_bytes=12_000)
        assert image_marker in studio_pdf_content, (
            f"Expected at least one embedded chart image object for chart_ids={studio_chart_ids}.\n"
            f"render exceptions:\n{caplog.text or 'No export_render_service exceptions captured.'}"
        )

        for chart_id in studio_chart_ids:
            single_chart_pdf = _assert_pdf_export_for_chart_ids([chart_id], minimum_bytes=8_000)
            assert image_marker in single_chart_pdf, (
                f"Expected embedded image object for chart_id={chart_id}, but none was found.\n"
                f"render exceptions:\n{caplog.text or 'No export_render_service exceptions captured.'}"
            )

        report = client.get(f"/report/{dataset_id}")
        assert report.status_code == 200
        report_body = report.json()
        assert report_body["executive_summary"]["insight"]
        assert report_body["branding"]["report_title"] == "Executive Decision Intelligence Report"
        assert report_body["business_story"]["business_story"]
        assert report_body["analysis_guardrails"]["invalid_methods"]
        assert report_body["data_quality_score"]["grade"]
        assert report_body["suggested_questions"]
        report_chart = next((chart for chart in report_body["chart_specs"] if chart["chart_id"] == visual_body["chart"]["chart_id"]), None)
        assert report_chart is not None
        assert report_chart["title"] == visual_body["chart"]["title"]
        assert report_chart["chart_type"] == visual_body["chart"]["chart_type"]
        assert len([chart for chart in report_body["chart_specs"] if chart["chart_id"] == visual_body["chart"]["chart_id"]]) == 1

        csv_export = client.get(f"/report/{dataset_id}/export", params={"format": "csv"})
        assert csv_export.status_code == 200
        assert "region,sales" in csv_export.text
        pdf_export = client.get(f"/report/{dataset_id}/export", params={"format": "pdf"})
        assert pdf_export.status_code == 200
        assert pdf_export.content.startswith(b"%PDF")
        pptx_export = client.get(f"/report/{dataset_id}/export", params={"format": "pptx"})
        assert pptx_export.status_code == 200
        assert pptx_export.content.startswith(b"PK")
        xlsx_export = client.get(f"/report/{dataset_id}/export", params={"format": "xlsx"})
        assert xlsx_export.status_code == 200
        assert xlsx_export.content.startswith(b"PK")
        png_export = client.get(f"/report/{dataset_id}/export", params={"format": "png"})
        assert png_export.status_code == 200
        assert png_export.content.startswith(b"\x89PNG")
        selected_pdf = client.get(
            f"/report/{dataset_id}/export",
            params={"format": "pdf", "chart_ids": dashboard_body["chart_specs"][0]["chart_id"], "package": "board"},
        )
        assert selected_pdf.status_code == 200

        themes = client.get("/themes")
        assert themes.status_code == 200
        assert len(themes.json()["themes"]) >= 12

        theme_switch = client.post("/themes/active/executive_dark")
        assert theme_switch.status_code == 200
        assert theme_switch.json()["name"] == "executive_dark"
        assert settings.THEME_STATE_FILE.exists()
        assert "executive_dark" in settings.THEME_STATE_FILE.read_text(encoding="utf-8")
        themed_dashboard = client.get(f"/analytics/{dataset_id}/dashboard")
        assert themed_dashboard.json()["theme"]["name"] == "executive_dark"
        assert themed_dashboard.json()["chart_specs"][0]["metadata"]["theme"] == "executive_dark"
        client.post("/themes/active/power_bi_professional")

        branding = client.get("/branding")
        assert branding.status_code == 200
        assert branding.json()["primary_color"] == "#118DFF"
        updated_branding = client.put(
            "/branding",
            json={
                "company_name": "Northwind Analytics",
                "report_title": "Quarterly Board Review",
                "primary_color": "#0055AA",
                "secondary_color": "#223344",
                "accent_color": "#FFAA00",
            },
        )
        assert updated_branding.status_code == 200
        assert updated_branding.json()["company_name"] == "Northwind Analytics"
        branded_dashboard = client.get(f"/analytics/{dataset_id}/dashboard")
        assert branded_dashboard.json()["branding"]["company_name"] == "Northwind Analytics"
        branded_report = client.get(f"/report/{dataset_id}")
        assert branded_report.json()["branding"]["report_title"] == "Quarterly Board Review"

        generated_sql = client.post(
            f"/sql-lab/{dataset_id}/generate",
            json={"prompt": "Show top 10 regions by sales"},
        )
        assert generated_sql.status_code == 200
        assert "GROUP BY" in generated_sql.json()["sql"]
        sql_result = client.post(
            f"/sql-lab/{dataset_id}/query",
            json={"sql": generated_sql.json()["sql"], "limit": 10},
        )
        assert sql_result.status_code == 200
        assert sql_result.json()["rows"][0]["region"] == "South"
        saved_sql = client.post(
            f"/sql-lab/{dataset_id}/save",
            json={"name": "Top regions", "sql": generated_sql.json()["sql"]},
        )
        assert saved_sql.status_code == 200
        sql_history = client.get(f"/sql-lab/{dataset_id}/history")
        assert sql_history.status_code == 200
        assert sql_history.json()["saved_queries"]

        domain = client.get(f"/intelligence/{dataset_id}/domain")
        assert domain.status_code == 200
        assert domain.json()["detection"]["domain"] == "Sales"
        assert domain.json()["root_causes"]

        regional = client.get(f"/intelligence/{dataset_id}/regional")
        assert regional.status_code == 200
        assert regional.json()["available"] is True
        assert regional.json()["regional_kpis"]
        assert regional.json()["map_charts"]

        dax = client.post(f"/dax/{dataset_id}/generate", json={"prompt": "Create Revenue YTD"})
        assert dax.status_code == 200
        dax_body = dax.json()
        assert dax_body["phase_1_intent"]["refined_business_question"]
        assert dax_body["phase_2_validation"]["valid_logic_check"]
        assert dax_body["phase_3_design"]["selected_analysis_type"] == "Trend Analysis"
        assert dax_body["phase_4_dax"]["dax_measures"] == "No DAX required for this analysis"
        assert dax_body["phase_5_visual"]["best_visual_type"]
        assert dax_body["phase_6_interpretation"]["business_interpretation"]
        assert dax_body["phase_7_decision"]["next_best_analysis_step"]
        assert len(dax_body["phase_8_export"]["slide_ready_insight_points"]) >= 3
        assert dax_body["analysis_type"] == "Trend Analysis"
        assert dax_body["data_logic_validation"]["is_valid"] is False
        assert dax_body["data_logic_validation"]["dax_allowed"] is False
        assert dax_body["dax_output"] == "No DAX required - analysis is structural."
        assert dax_body["measure_preview"]["measure_name"]
        assert dax_body["business_question_refined"]
        assert dax_body["dax_measure"]
        assert dax_body["business_meaning"]
        assert dax_body["best_visual"]
        assert dax_body["dashboard_placement"]["page"]
        assert dax_body["key_insight"]
        assert dax_body["next_best_question"]
        assert dax_body["recommended_visual_types"]
        assert dax_body["dashboard_integration_guidance"]
        assert dax_body["pdf_ppt_business_interpretation"]
        assert dax_body["executive_insight_summary"]
        optimized_dax = client.post(
            f"/dax/{dataset_id}/optimize",
            json={"dax": dax_body["dax"]},
        )
        assert optimized_dax.status_code == 200
        assert optimized_dax.json()["data_logic_validation"]["dax_allowed"] is False
        assert optimized_dax.json()["best_visual"]
        assert optimized_dax.json()["next_best_question"]
        dax_library = client.get(f"/dax/{dataset_id}/library")
        assert dax_library.status_code == 200
        assert dax_library.json()["measures"]
        assert not any("YTD" in measure["name"] for measure in dax_library.json()["measures"])
    finally:
        shutil.rmtree(test_root, ignore_errors=True)


def test_upload_accepts_excel_workbook(monkeypatch):
    test_root = Path("data") / "test_runs" / uuid4().hex
    monkeypatch.setattr(settings, "DATA_DIR", test_root)
    monkeypatch.setattr(settings, "UPLOADS_DIR", test_root / "uploads")
    monkeypatch.setattr(settings, "PROCESSED_DIR", test_root / "processed")
    monkeypatch.setattr(settings, "METADATA_DIR", test_root / "metadata")
    monkeypatch.setattr(settings, "DATASETS_DIR", test_root / "datasets")
    monkeypatch.setattr(settings, "SAMPLES_DIR", test_root / "samples")
    monkeypatch.setattr(settings, "BRAND_ASSETS_DIR", test_root / "branding")
    monkeypatch.setattr(settings, "DATASETS_METADATA_FILE", test_root / "metadata" / "datasets.json")
    monkeypatch.setattr(settings, "BRANDING_FILE", test_root / "metadata" / "branding.json")
    monkeypatch.setattr(settings, "SQL_QUERIES_FILE", test_root / "metadata" / "sql_queries.json")
    monkeypatch.setattr(settings, "THEME_STATE_FILE", test_root / "metadata" / "theme_state.json")
    ensure_data_directories()

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["region", "sales"])
    sheet.append(["North", 100])
    sheet.append(["South", 150])
    buffer = BytesIO()
    workbook.save(buffer)

    try:
        client = TestClient(app)
        response = client.post(
            "/upload",
            files={
                "file": (
                    "sales.xlsx",
                    buffer.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert response.status_code == 200
        dataset_id = response.json()["dataset_id"]
        overview = client.get(f"/datasets/{dataset_id}/overview")
        assert overview.status_code == 200
        assert overview.json()["row_count"] == 2
        assert overview.json()["column_groups"]["numeric"] == ["sales"]
    finally:
        shutil.rmtree(test_root, ignore_errors=True)


def test_registering_two_distinct_studio_charts_keeps_both_in_report(monkeypatch):
    test_root = Path("data") / "test_runs" / uuid4().hex
    monkeypatch.setattr(settings, "DATA_DIR", test_root)
    monkeypatch.setattr(settings, "UPLOADS_DIR", test_root / "uploads")
    monkeypatch.setattr(settings, "PROCESSED_DIR", test_root / "processed")
    monkeypatch.setattr(settings, "METADATA_DIR", test_root / "metadata")
    monkeypatch.setattr(settings, "DATASETS_DIR", test_root / "datasets")
    monkeypatch.setattr(settings, "SAMPLES_DIR", test_root / "samples")
    monkeypatch.setattr(settings, "BRAND_ASSETS_DIR", test_root / "branding")
    monkeypatch.setattr(settings, "DATASETS_METADATA_FILE", test_root / "metadata" / "datasets.json")
    monkeypatch.setattr(settings, "BRANDING_FILE", test_root / "metadata" / "branding.json")
    monkeypatch.setattr(settings, "SQL_QUERIES_FILE", test_root / "metadata" / "sql_queries.json")
    monkeypatch.setattr(settings, "THEME_STATE_FILE", test_root / "metadata" / "theme_state.json")
    ensure_data_directories()
    theme_manager.set_active("power_bi_professional")

    try:
        client = TestClient(app)
        upload = client.post(
            "/upload",
            files={
                "file": (
                    "sales_multi.csv",
                    b"region,segment,sales,profit\nNorth,Consumer,100,25\nSouth,Corporate,150,40\nEast,Consumer,90,20\n",
                    "text/csv",
                )
            },
        )
        assert upload.status_code == 200
        dataset_id = upload.json()["dataset_id"]

        first_visual = client.post(
            f"/visual-builder/{dataset_id}/render",
            json={
                "chart_type": "horizontal_bar",
                "dimension": "region",
                "measure": "sales",
                "aggregation": "sum",
                "sort": "descending",
                "title": "Sales by Region",
                "data_labels": True,
            },
        )
        assert first_visual.status_code == 200
        first_chart = first_visual.json()["chart"]

        second_visual = client.post(
            f"/visual-builder/{dataset_id}/render",
            json={
                "chart_type": "bar",
                "dimension": "segment",
                "measure": "profit",
                "aggregation": "sum",
                "sort": "descending",
                "title": "Profit by Segment",
                "data_labels": True,
            },
        )
        assert second_visual.status_code == 200
        second_chart = second_visual.json()["chart"]

        assert first_chart["chart_id"] != second_chart["chart_id"]

        first_register = client.post(f"/visual-builder/{dataset_id}/register", json=first_chart)
        second_register = client.post(f"/visual-builder/{dataset_id}/register", json=second_chart)
        assert first_register.status_code == 200
        assert second_register.status_code == 200
        assert first_register.json()["registered"] is True
        assert second_register.json()["registered"] is True

        report = client.get(f"/report/{dataset_id}")
        assert report.status_code == 200
        report_body = report.json()
        chart_specs = report_body["chart_specs"]

        report_first = next((chart for chart in chart_specs if chart["chart_id"] == first_chart["chart_id"]), None)
        report_second = next((chart for chart in chart_specs if chart["chart_id"] == second_chart["chart_id"]), None)
        assert report_first is not None
        assert report_second is not None
        assert report_first["title"] == "Sales by Region"
        assert report_first["chart_type"] == "horizontal_bar"
        assert report_second["title"] == "Profit by Segment"
        assert report_second["chart_type"] == "bar"
        assert len([chart for chart in chart_specs if chart["chart_id"] == first_chart["chart_id"]]) == 1
        assert len([chart for chart in chart_specs if chart["chart_id"] == second_chart["chart_id"]]) == 1
    finally:
        shutil.rmtree(test_root, ignore_errors=True)


def test_registering_bar_horizontal_and_table_studio_charts_exports_without_pdf_fallback(monkeypatch):
    test_root = Path("data") / "test_runs" / uuid4().hex
    monkeypatch.setattr(settings, "DATA_DIR", test_root)
    monkeypatch.setattr(settings, "UPLOADS_DIR", test_root / "uploads")
    monkeypatch.setattr(settings, "PROCESSED_DIR", test_root / "processed")
    monkeypatch.setattr(settings, "METADATA_DIR", test_root / "metadata")
    monkeypatch.setattr(settings, "DATASETS_DIR", test_root / "datasets")
    monkeypatch.setattr(settings, "SAMPLES_DIR", test_root / "samples")
    monkeypatch.setattr(settings, "BRAND_ASSETS_DIR", test_root / "branding")
    monkeypatch.setattr(settings, "DATASETS_METADATA_FILE", test_root / "metadata" / "datasets.json")
    monkeypatch.setattr(settings, "BRANDING_FILE", test_root / "metadata" / "branding.json")
    monkeypatch.setattr(settings, "SQL_QUERIES_FILE", test_root / "metadata" / "sql_queries.json")
    monkeypatch.setattr(settings, "THEME_STATE_FILE", test_root / "metadata" / "theme_state.json")
    ensure_data_directories()
    theme_manager.set_active("power_bi_professional")

    try:
        client = TestClient(app)
        upload = client.post(
            "/upload",
            files={
                "file": (
                    "studio_three_types.csv",
                    b"region,segment,sales,profit\nNorth,Consumer,100,25\nSouth,Corporate,150,40\nEast,Consumer,90,20\n",
                    "text/csv",
                )
            },
        )
        assert upload.status_code == 200
        dataset_id = upload.json()["dataset_id"]

        chart_requests = [
            {
                "chart_type": "bar",
                "dimension": "region",
                "measure": "sales",
                "aggregation": "sum",
                "sort": "descending",
                "title": "Studio Bar by Region",
                "data_labels": True,
            },
            {
                "chart_type": "horizontal_bar",
                "dimension": "segment",
                "measure": "sales",
                "aggregation": "sum",
                "sort": "descending",
                "title": "Studio Horizontal Bar by Segment",
                "data_labels": True,
            },
            {
                "chart_type": "table",
                "dimension": "segment",
                "measure": "profit",
                "aggregation": "sum",
                "sort": "descending",
                "title": "Studio Table by Segment",
                "data_labels": True,
            },
        ]

        created_charts: list[dict] = []
        for request in chart_requests:
            render_response = client.post(f"/visual-builder/{dataset_id}/render", json=request)
            assert render_response.status_code == 200
            chart = render_response.json()["chart"]
            register_response = client.post(f"/visual-builder/{dataset_id}/register", json=chart)
            assert register_response.status_code == 200
            created_charts.append(
                {
                    "chart_id": chart["chart_id"],
                    "title": chart["title"],
                    "chart_type": chart["chart_type"],
                }
            )

        report = client.get(f"/report/{dataset_id}")
        assert report.status_code == 200
        report_charts = report.json()["chart_specs"]
        for expected in created_charts:
            matched = next((chart for chart in report_charts if chart["chart_id"] == expected["chart_id"]), None)
            assert matched is not None
            assert matched["title"] == expected["title"]
            assert matched["chart_type"] == expected["chart_type"]

        pdf_export = client.get(
            f"/report/{dataset_id}/export",
            params=[("format", "pdf"), ("package", "board")] + [("chart_ids", chart["chart_id"]) for chart in created_charts],
        )
        assert pdf_export.status_code == 200
        assert pdf_export.content.startswith(b"%PDF")
        assert b"Chart image could not be rendered" not in pdf_export.content
    finally:
        shutil.rmtree(test_root, ignore_errors=True)


def test_studio_storyboard_add_assigns_order_and_prevents_duplicate(monkeypatch):
    test_root = Path("data") / "test_runs" / uuid4().hex
    monkeypatch.setattr(settings, "DATA_DIR", test_root)
    monkeypatch.setattr(settings, "UPLOADS_DIR", test_root / "uploads")
    monkeypatch.setattr(settings, "PROCESSED_DIR", test_root / "processed")
    monkeypatch.setattr(settings, "METADATA_DIR", test_root / "metadata")
    monkeypatch.setattr(settings, "DATASETS_DIR", test_root / "datasets")
    monkeypatch.setattr(settings, "SAMPLES_DIR", test_root / "samples")
    monkeypatch.setattr(settings, "BRAND_ASSETS_DIR", test_root / "branding")
    monkeypatch.setattr(settings, "DATASETS_METADATA_FILE", test_root / "metadata" / "datasets.json")
    monkeypatch.setattr(settings, "BRANDING_FILE", test_root / "metadata" / "branding.json")
    monkeypatch.setattr(settings, "SQL_QUERIES_FILE", test_root / "metadata" / "sql_queries.json")
    monkeypatch.setattr(settings, "THEME_STATE_FILE", test_root / "metadata" / "theme_state.json")
    ensure_data_directories()
    theme_manager.set_active("power_bi_professional")

    try:
        client = TestClient(app)
        upload = client.post(
            "/upload",
            files={
                "file": (
                    "storyboard_chart.csv",
                    b"region,sales\nNorth,100\nSouth,150\n",
                    "text/csv",
                )
            },
        )
        assert upload.status_code == 200
        dataset_id = upload.json()["dataset_id"]

        visual = client.post(
            f"/visual-builder/{dataset_id}/render",
            json={
                "chart_type": "horizontal_bar",
                "dimension": "region",
                "measure": "sales",
                "aggregation": "sum",
                "sort": "descending",
                "title": "Storyboard Sales by Region",
                "data_labels": True,
            },
        )
        assert visual.status_code == 200
        visual_body = visual.json()
        chart = visual_body["chart"]
        applied_spec = visual_body["applied_spec"]

        register = client.post(f"/visual-builder/{dataset_id}/register", json=chart)
        assert register.status_code == 200
        assert register.json()["chart"]["chart_id"] == chart["chart_id"]

        storyboard: list[dict] = []
        storyboard_entry = {
            "chart_id": chart["chart_id"],
            "title": chart["title"],
            "business_meaning": chart.get("metadata", {}).get("short_ai_insight", ""),
            "suggested_chart_type": applied_spec.get("chart_type", ""),
            "fields_used": chart.get("fields", []),
            "spec": applied_spec,
            "short_ai_insight": chart.get("metadata", {}).get("short_ai_insight", ""),
        }

        first_add = add_storyboard_entry(storyboard, storyboard_entry)
        assert first_add["added"] is True
        assert len(storyboard) == 1
        assert storyboard[0]["chart_id"] == chart["chart_id"]
        assert storyboard[0]["sequence"] == 1
        assert storyboard[0]["order"] == 1

        second_add = add_storyboard_entry(storyboard, storyboard_entry)
        assert second_add["added"] is False
        assert len(storyboard) == 1
        # Current behavior: duplicate adds return existing entry and do not update order/sequence.
        assert storyboard[0]["sequence"] == 1
        assert storyboard[0]["order"] == 1
    finally:
        shutil.rmtree(test_root, ignore_errors=True)


def test_data_cleaning_endpoint_produces_non_destructive_downloads(monkeypatch):
    test_root = Path("data") / "test_runs" / uuid4().hex
    monkeypatch.setattr(settings, "DATA_DIR", test_root)
    monkeypatch.setattr(settings, "UPLOADS_DIR", test_root / "uploads")
    monkeypatch.setattr(settings, "PROCESSED_DIR", test_root / "processed")
    monkeypatch.setattr(settings, "METADATA_DIR", test_root / "metadata")
    monkeypatch.setattr(settings, "DATASETS_DIR", test_root / "datasets")
    monkeypatch.setattr(settings, "SAMPLES_DIR", test_root / "samples")
    monkeypatch.setattr(settings, "BRAND_ASSETS_DIR", test_root / "branding")
    monkeypatch.setattr(settings, "DATASETS_METADATA_FILE", test_root / "metadata" / "datasets.json")
    monkeypatch.setattr(settings, "BRANDING_FILE", test_root / "metadata" / "branding.json")
    monkeypatch.setattr(settings, "SQL_QUERIES_FILE", test_root / "metadata" / "sql_queries.json")
    monkeypatch.setattr(settings, "THEME_STATE_FILE", test_root / "metadata" / "theme_state.json")
    ensure_data_directories()

    try:
        client = TestClient(app)
        upload = client.post(
            "/upload",
            files={"file": ("dirty.csv", b"region,sales,date\n North ,100,2024-01-01\nNorth,,2024-01-02\nnorth,5000,\n", "text/csv")},
        )
        assert upload.status_code == 200
        dataset_id = upload.json()["dataset_id"]

        clean = client.post(
            f"/datasets/{dataset_id}/clean",
            json={
                "normalize_casing": "lower",
                "numeric_missing_strategy": "median",
                "categorical_missing_strategy": "mode",
                "datetime_missing_strategy": "manual",
                "outlier_strategy": "cap",
                "outlier_method": "iqr",
            },
        )
        assert clean.status_code == 200
        body = clean.json()
        assert body["rows_after"] <= body["rows_before"]
        assert body["completeness_after_pct"] >= body["completeness_before_pct"]
        assert body["cleaned_filename_csv"].endswith(".csv")
        assert body["cleaned_filename_xlsx"].endswith(".xlsx")

        csv_download = client.get(f"/datasets/{dataset_id}/clean/download", params={"filename": body["cleaned_filename_csv"]})
        xlsx_download = client.get(f"/datasets/{dataset_id}/clean/download", params={"filename": body["cleaned_filename_xlsx"]})
        assert csv_download.status_code == 200
        assert xlsx_download.status_code == 200
        assert b"region,sales,date" in csv_download.content
        assert xlsx_download.content.startswith(b"PK")
    finally:
        shutil.rmtree(test_root, ignore_errors=True)


def test_excel_export_contains_dashboard_summary_raw_data_schema_and_chart_object(monkeypatch):
    test_root = Path("data") / "test_runs" / uuid4().hex
    monkeypatch.setattr(settings, "DATA_DIR", test_root)
    monkeypatch.setattr(settings, "UPLOADS_DIR", test_root / "uploads")
    monkeypatch.setattr(settings, "PROCESSED_DIR", test_root / "processed")
    monkeypatch.setattr(settings, "METADATA_DIR", test_root / "metadata")
    monkeypatch.setattr(settings, "DATASETS_DIR", test_root / "datasets")
    monkeypatch.setattr(settings, "SAMPLES_DIR", test_root / "samples")
    monkeypatch.setattr(settings, "BRAND_ASSETS_DIR", test_root / "branding")
    monkeypatch.setattr(settings, "DATASETS_METADATA_FILE", test_root / "metadata" / "datasets.json")
    monkeypatch.setattr(settings, "BRANDING_FILE", test_root / "metadata" / "branding.json")
    monkeypatch.setattr(settings, "SQL_QUERIES_FILE", test_root / "metadata" / "sql_queries.json")
    monkeypatch.setattr(settings, "THEME_STATE_FILE", test_root / "metadata" / "theme_state.json")
    ensure_data_directories()
    theme_manager.set_active("power_bi_professional")

    try:
        client = TestClient(app)
        upload = client.post(
            "/upload",
            files={"file": ("sales_chart.csv", b"region,sales\nNorth,100\nSouth,150\nEast,90\n", "text/csv")},
        )
        assert upload.status_code == 200
        dataset_id = upload.json()["dataset_id"]
        xlsx_export = client.get(f"/report/{dataset_id}/export", params={"format": "xlsx"})
        assert xlsx_export.status_code == 200
        workbook = load_workbook(BytesIO(xlsx_export.content))
        assert "Dashboard Summary" in workbook.sheetnames
        assert "Raw Data" in workbook.sheetnames
        assert "Column Schema" in workbook.sheetnames
        assert len(workbook["Dashboard Summary"]._charts) >= 1
    finally:
        shutil.rmtree(test_root, ignore_errors=True)
