from io import BytesIO
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook
from pptx import Presentation

from backend.services.excel_export_service import build_executive_excel
from backend.services.pdf_export_service import build_executive_pdf
from backend.services.ppt_export_service import build_executive_pptx


def _sample_report():
    chart = {
        "chart_id": "story_chart_revenue",
        "title": "Revenue by Segment",
        "chart_type": "bar",
        "columns": ["segment", "revenue"],
        "metadata": {"short_ai_insight": "Segment B leads validated revenue.", "top_categories": {"segment": [{"label": "A", "value": 100}, {"label": "B", "value": 180}]}},
        "plotly": {},
    }
    kpi = {"kpi_id": "total_revenue", "label": "Total Revenue", "value": 280, "status": "positive", "business_context": "Validated KPI from existing engine."}
    ai_card = {
        "type": "Opportunity",
        "title": "Opportunity: Focus on Revenue",
        "business_meaning": "Revenue is the strongest validated KPI.",
        "supporting_evidence": "KPI discovery returned revenue.",
        "expected_business_impact": "Better focus on revenue drivers.",
        "executive_recommendation": "Compare revenue by segment.",
        "confidence_score": 0.8,
    }
    rec = {
        "title": "Opportunity: Focus on Revenue",
        "priority": "High",
        "business_value": "Revenue is the strongest validated KPI.",
        "difficulty": "Low",
        "expected_impact": "Better focus on revenue drivers.",
        "recommendation": "Compare revenue by segment.",
    }
    storyboard = {
        "sections": [
            {"section_id": "executive_summary", "content": {"dataset_readiness": {"score": 91, "ready": True, "reason": "Ready"}, "overall_business_health": 94, "executive_summary": "Dataset is ready for executive review.", "top_opportunity": "Revenue growth", "biggest_risk": "No major risk"}},
            {"section_id": "kpi_overview", "kpis": [kpi]},
            {"section_id": "ai_business_insights", "cards": [ai_card]},
            {"section_id": "executive_charts", "charts": [chart]},
            {"section_id": "executive_recommendations", "recommendations": [rec]},
        ],
        "source_payloads": {"data_insights_status": "ready", "ai_business_insights_status": "ready", "dashboard_status": "ready"},
    }
    return {
        "dataset_id": "sample_dataset",
        "branding": {"company_name": "AI Analytics", "report_title": "Executive Storyboard", "primary_color": "#118DFF", "accent_color": "#E66C37"},
        "overview": {"row_count": 2, "column_count": 2},
        "kpi_cards": [kpi],
        "chart_specs": [chart],
        "chart_count": 1,
        "executive_summary": {"insight": "Ready", "reason": "Validated", "action": "Proceed"},
        "business_story": {"business_story": "Validated storyboard."},
        "data_quality_score": {"score": 94, "grade": "A", "completeness_pct": 100, "duplicate_pct": 0, "explanation": "Clean"},
        "executive_storyboard": storyboard,
    }


def test_storyboard_pdf_export_is_valid():
    data = build_executive_pdf(_sample_report(), package="storyboard")

    assert data.startswith(b"%PDF")
    assert len(data) > 1000
    assert b"/Subtype /Image" in data or b"/Subtype/Image" in data


def test_storyboard_pptx_export_opens_programmatically():
    data = build_executive_pptx(_sample_report(), package="storyboard")
    deck = Presentation(BytesIO(data))

    assert len(deck.slides) >= 5
    with ZipFile(BytesIO(data)) as archive:
        assert any(name.startswith("ppt/media/image") for name in archive.namelist())


def test_storyboard_excel_export_contains_summary_sheet():
    raw_df = pd.DataFrame({"segment": ["A", "B"], "revenue": [100, 180]})
    data = build_executive_excel(_sample_report(), raw_df=raw_df, summary={"missing_values_by_column": {}, "dtypes": {}}, package="storyboard")
    workbook = load_workbook(BytesIO(data))

    assert "Storyboard Summary" in workbook.sheetnames
    sheet = workbook["Storyboard Summary"]
    assert sheet["A1"].value == "Executive Storyboard Summary"
    assert "Visual Dashboard" in workbook.sheetnames
    assert len(workbook["Visual Dashboard"]._images) >= 1
    assert "Pivot Tables" in workbook.sheetnames
    pivot_sheet = workbook["Pivot Tables"]
    assert len(pivot_sheet._charts) >= 1
